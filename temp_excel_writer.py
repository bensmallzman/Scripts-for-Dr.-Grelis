import openpyxl
import sys

location = str(sys.argv[1]) + ".xlsx"

try:
    wb = openpyxl.load_workbook(location)
except:
    wb = openpyxl.Workbook()
sheet = wb.active
c1 = sheet.cell(row = 1, column = 1)
c1.value = "s/n"
c2 = sheet.cell(row = 1, column = 2)
c2.value = "unit"
c3 = sheet.cell(row = 1, column = 3)
c3.value = "single/multi"
c4 = sheet.cell(row = 1, column = 4)
c4.value = "measured"
c5 = sheet.cell(row = 1, column = 5)
c5.value = "measured"
c6 = sheet.cell(row = 1, column = 6)
c6.value = "measured"
c7 = sheet.cell(row = 1, column = 7)
c7.value = "final"
c8 = sheet.cell(row = 1, column = 8)
c8.value = "pass(Y/N)"
c9 = sheet.cell(row = 1, column = 9)
c9.value = "Tech Initials"
c10 = sheet.cell(row = 1, column = 10)
c10.value = "Tech Signature"

f = open("data.txt", 'r')

serial = ''
vol = ''
channel = ''
measured0 = ''
measured1 = ''
measured2 = ''
average = ''

for row in f.readlines():
    if (row.startswith("Pipette")):
        serial = row.partition(": ")[2]
    if (row.startswith("Volume")):
        vol = row.partition(": ")[2]
    if (row.startswith("Channel")):
        channel = row.partition(": ")[2]
    if (row.startswith("Example volumes")):
        x = row.partition(": ")[2]
        temp = x.split()
        measured0 = temp[0]
        measured1 = temp[1]
        measured2 = temp[2]
    if (row.startswith("Example Average")):
        average = row.partition(": ")[2]
    tup = (serial, vol, channel, measured0, measured1, measured2, average)
    if (all(tup)):
        sheet.append(tup)
        serial, vol, channel, measured0, measured1, measured2, average = None, None, None, None, None, None, None


wb.save(filename = location)
